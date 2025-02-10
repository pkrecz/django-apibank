# -*- coding: utf-8 -*-

import datetime
import decimal
from rest_framework import viewsets
from rest_framework import status
from rest_framework.decorators import action
from rest_framework.exceptions import APIException
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django.http import JsonResponse
from django.http.response import HttpResponse
from django.db import transaction
from django.db.models import ProtectedError
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font
from drf_yasg.utils import swagger_auto_schema
from .models import CustomerModel, AccountModel, AccountTypeModel, ParameterModel, OperationModel, LogModel
from .serializers import (
                            CustomerCreateSerializer, CustomerUpdateSerializer, CustomerLRDSerializer,
                            AccountCreateSerializer, AccountUpdateSerializer, AccountLRDSerializer, AccountOperationSerializer,
                            AccountTypeCLRDSerializer, AccountTypeUpdateSerializer,
                            ParameterSerializer,
                            OperationNewSerializer, OperationHistorySerializer, OperationInterestSerializer,
                            LogMonitoringSerializer)
from .decorators import ActivityMonitoringClass
from .filters import CustomerFilter, AccountFilter, AccountTypeFilter, LogFilter
from .functions import generate_iban, get_balance_and_debit


""" Customer """
class CustomerViewSet(viewsets.ModelViewSet):
    http_method_names = ["get", "post", "put", "delete"]
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    queryset = CustomerModel.objects.all()
    filterset_class = CustomerFilter
    search_fields = ["pesel", "identification"]
    ordering_fields = ["last_name", "first_name"]
    ordering = ["last_name"]
    swagger_viewset_tag = ["Customer"]

    def get_serializer_class(self):
        match self.action:
            case "create":
                return CustomerCreateSerializer
            case "update":
                return CustomerUpdateSerializer
            case _:
                return CustomerLRDSerializer

    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    def create(self, request, *args, **kwargs):
        try:
            serializer = self.get_serializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            instance = serializer.save(created_employee=self.request.user)
            return_serializer = CustomerLRDSerializer(instance, context={"request": request})
            return JsonResponse(data=return_serializer.data, status=status.HTTP_201_CREATED)
        except APIException as exc:
            return JsonResponse(data=exc.detail, status=status.HTTP_400_BAD_REQUEST)

    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        try:
            self.perform_destroy(instance)
            return JsonResponse(data={"message": "Customer has been deleted."}, status=status.HTTP_200_OK)
        except ProtectedError:
            return JsonResponse(data={"message": "Deletion impossible. This record has referenced data!"}, status=status.HTTP_400_BAD_REQUEST)
        except APIException as exc:
            return JsonResponse(data=exc.detail, status=status.HTTP_400_BAD_REQUEST)


""" Account """
class AccountViewSet(viewsets.ModelViewSet):
    http_method_names = ["get", "post", "put", "delete"]
    queryset = AccountModel.objects.all()
    filterset_class = AccountFilter
    search_fields = ["number_iban"]
    ordering_fields = ["number_iban"]
    swagger_viewset_tag = ["Account"]

    def get_serializer_class(self):
        match self.action:
            case "newoperation":
                return OperationNewSerializer
            case "create":
                return AccountCreateSerializer
            case "update":
                return AccountUpdateSerializer
            case _:
                return AccountLRDSerializer

    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    @ActivityMonitoringClass()
    def create(self, request, *args, **kwargs):
        try:
            serializer = self.get_serializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            serializer.validated_data["free_balance"] = serializer.validated_data.get("debit")
            serializer.fields["free_balance"].read_only = False
            instance = serializer.save(created_employee=self.request.user)
            return_serializer = AccountLRDSerializer(instance, context={"request": request})
            return JsonResponse(data=return_serializer.data, status=status.HTTP_201_CREATED)
        except APIException as exc:
            return JsonResponse(data=exc.detail, status=status.HTTP_400_BAD_REQUEST)

    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)

    @ActivityMonitoringClass()
    def update(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            serializer = self.get_serializer(instance, data=request.data)
            serializer.is_valid(raise_exception=True)
            free_balance = instance.balance + serializer.validated_data.get("debit")
            serializer.validated_data["free_balance"] = free_balance
            serializer.fields["free_balance"].read_only = False
            instance = serializer.save()
            return_serializer = AccountLRDSerializer(instance, context={"request": request})
            return JsonResponse(data=return_serializer.data, status=status.HTTP_200_OK)
        except APIException as exc:
            return JsonResponse(data=exc.detail, status=status.HTTP_400_BAD_REQUEST)

    @ActivityMonitoringClass()
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        try:
            self.perform_destroy(instance)
            return JsonResponse(data={"message": "Account has been deleted."}, status=status.HTTP_200_OK)
        except ProtectedError:
            return JsonResponse(data={"message": "Deletion impossible. This record has referenced data!"}, status=status.HTTP_400_BAD_REQUEST)
        except APIException as exc:
            return JsonResponse(data=exc.detail, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=["post"])
    @ActivityMonitoringClass()
    def generate(self, request, pk=None):
        instance = self.get_object()
        if not instance.number_iban:
            try:
                account = instance.id_account
                customer = instance.customer_id
                country_code = ParameterModel.objects.get().country_code
                bank_number = ParameterModel.objects.get().bank_number
                subaccount = AccountTypeModel.objects.get(id_account_type=instance.account_type_id).subaccount
                iban = generate_iban(
                                        country_code=country_code,
                                        bank_number=bank_number,
                                        subaccount=subaccount,
                                        account=str(account),
                                        customer=str(customer))
                serializer = AccountLRDSerializer(instance, data={"number_iban": iban}, partial=True)
                serializer.is_valid(raise_exception=True)
                instance = serializer.save()
                return_serializer = AccountLRDSerializer(instance, context={"request": request})
                return JsonResponse(data=return_serializer.data, status=status.HTTP_200_OK)
            except APIException as exc:
                return JsonResponse(data=exc.detail, status=status.HTTP_400_BAD_REQUEST)
        else:
            return JsonResponse({"message": "IBAN number already exists."}, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"])
    @ActivityMonitoringClass()
    def newoperation(self, request, pk=None):
        instance = self.get_object()
        if request.method == "POST":
            try:
                with transaction.atomic():
                    serializer_operation = OperationNewSerializer(data=request.data)
                    serializer_operation.is_valid(raise_exception=True)
                    # Choosing operation type
                    if serializer_operation.validated_data.get("type_operation") == 2:
                        serializer_operation.validated_data["value_operation"] *= (-1)
                    # Getting data from account
                    balance, debit = get_balance_and_debit(instance.id_account)
                    # Calculating balance & free_balance after transaction
                    balance_after_operation = balance + serializer_operation.validated_data.get("value_operation")
                    free_balance_after_operation = balance_after_operation + debit
                    # Updating balance & free balance for account
                    data_account = {
                                        "balance": balance_after_operation,
                                        "free_balance": free_balance_after_operation}
                    serializer_account = AccountOperationSerializer(instance, data=data_account, partial=True)
                    serializer_account.is_valid(raise_exception=True)
                    instance = serializer_account.save()
                    # Creating new operation
                    serializer_operation.validated_data["id_account"] = instance
                    serializer_operation.validated_data["balance_after_operation"] = balance_after_operation
                    serializer_operation.save(operation_employee=self.request.user)
            except APIException as exc:
                return JsonResponse(data=exc.detail, status=status.HTTP_400_BAD_REQUEST)
        return_serializer = AccountLRDSerializer(instance, context={"request": request})
        return JsonResponse(data=return_serializer.data, status=status.HTTP_200_OK)

    @action(detail=True, methods=["get"])
    def operations(self, request, pk=None):
        instance = self.get_object()
        try:
            queryset = OperationModel.objects.filter(id_account=instance.id_account).order_by("-operation_date")
            queryset = self.filter_queryset(queryset)
            page = self.paginate_queryset(queryset)
            if page is not None:
                serializer = OperationHistorySerializer(page, context={"request": request}, many=True)
                return self.get_paginated_response(data=serializer.data)
            serializer = OperationHistorySerializer(queryset, context={"request": request}, many=True)
            return JsonResponse(data=serializer.data, safe=False, status=status.HTTP_200_OK)
        except APIException as exc:
                return JsonResponse(data=exc.detail, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=["get"])
    def export(self, request, pk=None):
        side = Side(style="dashed", color="FF000000")
        border_around = Border(left=side, right=side, top=side, bottom=side)
        file_name = "History_operations.xlsx"
        fields = [
                    "id_operation",
                    "type_operation",
                    "value_operation",
                    "balance_after_operation",
                    "operation_date"]
        instance = self.get_object()
        data = OperationModel.objects.filter(id_account=instance.id_account).order_by("-operation_date").values_list(*fields)
        data = self.filter_queryset(data)
        if not data.exists():
            return JsonResponse({"message": "No data to be exported"}, status=status.HTTP_400_BAD_REQUEST)
        response = HttpResponse(content_type="application/vnd.ms-excel")
        response["Content-Disposition"] = f"attachment; filename={file_name}"
        workbook = Workbook()
        workbook.iso_dates = True
        worksheet = workbook.active
        worksheet.title = "Operations"
        # Column headers
        headers = [
                    "Id operation",
                    "Typ of operation",
                    "Value operation",
                    "Balance after operation",
                    "Operation date"]
        for column_number, column_title in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=column_number)
            cell.value = column_title
            cell.font = Font(bold=True, italic=True)
            cell.fill = PatternFill(fgColor="0000FFFF", fill_type="solid")
            cell.border = border_around
        # Cell data
        for row_number, row in enumerate(data, 1):
            for column_number, cell_value in enumerate(row, 1):
                if type(cell_value) is datetime.datetime:
                    cell_value = cell_value.replace(tzinfo=None)
                    cell_value = cell_value.strftime('%d.%m.%Y %H:%M:%S')
                if column_number == 2:
                    match cell_value:
                        case 1:
                            cell_value = "Deposit"
                        case 2:
                            cell_value = "Withdrawal"
                        case 3:
                            cell_value = "Interest"
                cell = worksheet.cell(row=row_number+1, column=column_number)
                cell.value = cell_value
                cell.border = border_around
                if type(cell_value) is decimal.Decimal:
                    cell.number_format = "#,##0.00"
        # AutoFit column width
        for column in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            adjusted_width = (max_length + 2) * 1.1
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
        workbook.save(response)
        return response

    @action(detail=False, methods=["post"])
    @ActivityMonitoringClass()
    def interest(self, request, pk=None):
        data = AccountModel.objects.filter(balance__gt = 0, percent__gt = 0)
        if data.exists():
            try:
                with transaction.atomic():
                    counter = 0
                    for instance in data:
                        # Calculating data
                        interest = round(instance.balance * (instance.percent / 100), 2)
                        new_balance = instance.balance + interest
                        new_free_balance = new_balance + instance.debit
                        # Updating balance & free balance for account
                        data_account = {
                                            "balance": new_balance,
                                            "free_balance": new_free_balance}
                        serializer = AccountOperationSerializer(instance, data=data_account, partial=True)
                        serializer.is_valid(raise_exception=True)
                        serializer.save()
                        # Creating new operation
                        data_operation = {
                                            "type_operation": 3,
                                            "value_operation": interest,
                                            "balance_after_operation": new_balance,
                                            "id_account": instance.id_account}
                        serializer = OperationInterestSerializer(data=data_operation)
                        serializer.is_valid(raise_exception=True)
                        serializer.save(operation_employee=self.request.user)
                        counter += 1
                    msg = f"Interest for {counter} account(s) has been recounted."
                    return JsonResponse(data={"message": msg}, status=status.HTTP_200_OK)
            except APIException as exc:
                return JsonResponse(data=exc.detail, status=status.HTTP_400_BAD_REQUEST)
        else:
            return JsonResponse(data={"message": "No accounts to be recounted."}, status=status.HTTP_200_OK)


""" Account Type """
class AccountTypeViewSet(viewsets.ModelViewSet):
    http_method_names = ["get", "post", "put", "delete"]
    queryset = AccountTypeModel.objects.all()
    filterset_class = AccountTypeFilter
    search_fields = ["code", "description"]
    ordering_fields = ["code"]
    swagger_viewset_tag = ["Account Type"]

    def get_serializer_class(self):
        match self.action:
            case "update":
                return AccountTypeUpdateSerializer
            case _:
                return AccountTypeCLRDSerializer

    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)
    
    def create(self, request, *args, **kwargs):
        return super().create(request, *args, **kwargs)
    
    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)
    
    def update(self, request, *args, **kwargs):
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        try:
            self.perform_destroy(instance)
            return JsonResponse(data={"message": "Account type has been deleted."}, status=status.HTTP_200_OK)
        except ProtectedError:
            return JsonResponse(data={"message": "Deletion impossible. This record has referenced data!"}, status=status.HTTP_400_BAD_REQUEST)


""" Parameter """
class ParameterViewSet(viewsets.ModelViewSet):
    http_method_names = ["get", "post", "put"]
    queryset = ParameterModel.objects.all()
    serializer_class = ParameterSerializer
    swagger_viewset_tag = ["Parameter"]

    @swagger_auto_schema(auto_schema=None)
    def list(self, request, *args, **kwargs):
        return JsonResponse(data={"message": "Method is not allowed."}, status=status.HTTP_405_METHOD_NOT_ALLOWED)

    def create(self, request, *args, **kwargs):
        if ParameterModel.objects.count() != 0:
            return JsonResponse(data={"message": "Data already exists."}, status=status.HTTP_400_BAD_REQUEST)
        return super().create(request, *args, **kwargs)

    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)
    
    def update(self, request, *args, **kwargs):
        return super().update(request, *args, **kwargs)


""" Log """
class LogViewSet(viewsets.ModelViewSet):
    http_method_names = ["get"]
    queryset = LogModel.objects.all().order_by("-date_log")
    serializer_class = LogMonitoringSerializer
    filterset_class = LogFilter
    search_fields = ["user_log"]
    ordering_fields = ["date_log", "duration_log"]
    ordering = ["-duration_log"]
    swagger_viewset_tag = ["Log"]

    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    @swagger_auto_schema(auto_schema=None)
    def retrieve(self, request, *args, **kwargs):
        return JsonResponse(data={"message": "Method is not allowed."}, status=status.HTTP_405_METHOD_NOT_ALLOWED)
